from __future__ import annotations

import json
import shutil
import argparse
from dataclasses import dataclass, asdict
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import load_workbook


ROOT = Path(r"D:\FISH\fish-sync")
PHOTO_WORKBOOK = ROOT / "public" / "horoshop_photo_import.xlsx"
OUTPUT_ROOT = ROOT / "public" / "horoshop-image-utility"
REPORT_PATH = ROOT / "data" / "horoshop_image_utility_report.json"
DEFAULT_PUBLIC_BASE_URL = "https://vsedliarybalky.com.ua"
MAX_FILES_PER_BATCH = 500
MAX_BYTES_PER_BATCH = 240 * 1024 * 1024


@dataclass
class BatchStats:
    name: str
    file_count: int = 0
    total_bytes: int = 0


def local_path_from_url(url: str, public_base_url: str) -> Path:
    expected = urlparse(public_base_url.rstrip("/") + "/")
    parsed = urlparse(url)
    if parsed.netloc and parsed.netloc != expected.netloc:
        raise ValueError(f"Unexpected host in URL: {url}")

    rel = unquote(parsed.path.lstrip("/") if parsed.scheme else url.lstrip("/"))
    if not rel.startswith("photo-import/"):
        raise ValueError(f"Unexpected image path in URL: {url}")
    return ROOT / "public" / Path(*rel.split("/"))


def next_gallery_name(article: str, idx: int, ext: str) -> str:
    if idx == 1:
        return f"{article}@gallery_common{ext}"
    return f"{article}@gallery_common@{idx}{ext}"


def ensure_batch(
    output_root: Path,
    current_batch: Path,
    stats: BatchStats,
    next_size: int,
) -> tuple[Path, BatchStats]:
    if (
        stats.file_count < MAX_FILES_PER_BATCH
        and stats.total_bytes + next_size <= MAX_BYTES_PER_BATCH
    ):
        return current_batch, stats

    batch_index = int(current_batch.name.split("-")[-1]) + 1
    next_batch = output_root / f"batch-{batch_index:03d}"
    next_batch.mkdir(parents=True, exist_ok=True)
    return next_batch, BatchStats(name=next_batch.name)


def build_utility(
    workbook: Path,
    output_root: Path,
    report_path: Path,
    public_base_url: str,
) -> dict:
    output_root.mkdir(parents=True, exist_ok=True)
    for child in output_root.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()

    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = [str(h or "").strip() for h in rows[0]]
    article_idx = header.index("Артикул")
    gallery_idx = header.index("Галерея")

    batch = output_root / "batch-001"
    batch.mkdir(parents=True, exist_ok=True)
    stats = BatchStats(name=batch.name)

    total_articles = 0
    total_files = 0
    missing_sources: list[dict[str, str]] = []
    batches: list[BatchStats] = []

    for row in rows[1:]:
        article = str(row[article_idx] or "").strip()
        gallery = str(row[gallery_idx] or "").strip()
        if not article or not gallery:
            continue

        urls = [u.strip() for u in gallery.split(";") if u.strip()]
        if not urls:
            continue

        total_articles += 1
        for idx, url in enumerate(urls, start=1):
            source = local_path_from_url(url, public_base_url)
            if not source.exists():
                missing_sources.append({"article": article, "url": url})
                continue

            next_batch, next_stats = ensure_batch(
                output_root,
                batch,
                stats,
                source.stat().st_size,
            )
            if next_batch != batch:
                batches.append(stats)
                batch, stats = next_batch, next_stats
            target_name = next_gallery_name(article, idx, source.suffix.lower())
            target = batch / target_name
            shutil.copy2(source, target)

            stats.file_count += 1
            stats.total_bytes += target.stat().st_size
            total_files += 1

    batches.append(stats)
    report = {
        "source_workbook": str(workbook),
        "output_root": str(output_root),
        "public_base_url": public_base_url,
        "articles_prepared": total_articles,
        "files_prepared": total_files,
        "batch_limit_files": MAX_FILES_PER_BATCH,
        "batch_limit_bytes": MAX_BYTES_PER_BATCH,
        "batches": [asdict(b) for b in batches],
        "missing_sources": missing_sources,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=PHOTO_WORKBOOK)
    parser.add_argument("--output-root", type=Path, default=OUTPUT_ROOT)
    parser.add_argument("--report", type=Path, default=REPORT_PATH)
    parser.add_argument("--public-base-url", default=DEFAULT_PUBLIC_BASE_URL)
    args = parser.parse_args()

    report = build_utility(
        workbook=args.workbook,
        output_root=args.output_root,
        report_path=args.report,
        public_base_url=args.public_base_url,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
