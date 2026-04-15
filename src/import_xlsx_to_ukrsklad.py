"""
Імпорт товарів з xlsx-експортів у УкрСклад7.

Сценарій:
  1. Читає xlsx-файли з папки або за списком шляхів
  2. Визначає цільову категорію та сімейство товару
  3. Створює відсутні групи TIP
  4. Upsert у TOVAR_NAME за артикулом (KOD)
  5. Upsert у TOVAR_ZAL для Мій Склад

За замовчуванням працює у dry-run.
"""
from __future__ import annotations

import argparse
import json
import os
from dataclasses import asdict, dataclass
from pathlib import Path

import fdb
from openpyxl import load_workbook

from catalog_rules import get_source_category, parse_product

ROOT = Path(r"D:\FISH\fish-sync")
PREVIEW_JSON = ROOT / "data" / "xlsx_import_preview.json"
FBCLIENT = ROOT / "tmp" / "fb3x64" / "fbclient.dll"
LIVE_DB = Path(r"C:\ProgramData\UkrSklad7\db\Sklad.tcb")
SNAPSHOT_DB = ROOT / "tmp" / "sklad_snapshot.fdb"

DEFAULT_HEADERS = [
    "Код",
    "Повна назва товару",
    "Од.вим.",
    "К-ть",
    "У резерві",
    "Ціна прих.",
    "Валюта приходу",
    "Розд. ціна",
    "Опт. ціна",
    "Ціна 1",
    "Ціна 2",
    "Валюта видаткова",
    "Гарантія",
    "Мін. залишок",
    "Послуга",
    "Фасовка",
    "К-ть в упаковці",
    "Додатково",
    "Фото",
    "Виробник",
    "Акція ціна",
]


@dataclass
class ImportRow:
    source_file: str
    code: str
    name: str
    unit: str
    qty: float
    purchase_price: float
    retail_price: float
    wholesale_price: float
    warranty: str
    min_qty: float
    packaging: float
    package_count: float
    extra: str
    manufacturer: str
    promo_price: float
    category_path: tuple[str, ...]
    family: str


def to_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def normalize_headers(row: tuple[object, ...]) -> list[str]:
    values = [str(cell).strip() if cell is not None else "" for cell in row]
    if values and values[0] == "Код":
        return values
    return DEFAULT_HEADERS


def iter_import_rows(path: Path) -> list[ImportRow]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    first_row = rows[0]
    if str(first_row[0] or "").strip() == "Код":
        headers = normalize_headers(first_row)
        data_rows = rows[1:]
    else:
        # У папці можуть лежати сторонні xlsx, які не є експортом УкрСкладу.
        # Якщо перший рядок не схожий на товарну позицію — просто пропускаємо файл.
        if len(first_row) < 4 or to_float(first_row[3]) == 0.0 and str(first_row[0] or "").strip().startswith("ID "):
            return []
        headers = DEFAULT_HEADERS
        data_rows = rows
    sample_names = [str(row[1]).strip() for row in data_rows[:5] if len(row) > 1 and row[1]]
    category_path, family = get_source_category(path.stem, sample_names)

    imported: dict[str, ImportRow] = {}
    for raw in data_rows:
        if not raw or raw[0] in (None, "", "Код"):
            continue
        row = list(raw) + [None] * max(0, len(headers) - len(raw))
        code = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] is not None else ""
        if not code or not name:
            continue
        imported[code] = ImportRow(
            source_file=path.name,
            code=code,
            name=name,
            unit=str(row[2] or "шт").strip() or "шт",
            qty=to_float(row[3]),
            purchase_price=to_float(row[5]),
            retail_price=to_float(row[7]),
            wholesale_price=to_float(row[8]),
            warranty=str(row[12] or "").strip(),
            min_qty=to_float(row[13]),
            packaging=to_float(row[15] or 1),
            package_count=to_float(row[16] or 1),
            extra=str(row[17] or "").strip(),
            manufacturer=str(row[19] or "").strip(),
            promo_price=to_float(row[20]),
            category_path=category_path,
            family=family,
        )
    return list(imported.values())


def connect(db_path: Path):
    os.environ["PATH"] = str(FBCLIENT.parent) + os.pathsep + os.environ.get("PATH", "")
    return fdb.connect(
        database=str(db_path),
        user="SYSDBA",
        password="masterkey",
        charset="UTF8",
        fb_library_name=str(FBCLIENT),
    )


def ensure_tip(conn, path: tuple[str, ...]) -> int:
    parent = 1
    cur = conn.cursor()
    for name in path:
        row = cur.execute(
            "SELECT NUM FROM TIP WHERE NAME = ? AND GRUPA = ? AND VISIBLE = 1",
            (name, parent),
        ).fetchone()
        if row:
            parent = int(row[0])
            continue
        cur.execute(
            """
            INSERT INTO TIP (NAME, GRUPA, VISIBLE)
            VALUES (?, ?, 1)
            RETURNING NUM
            """,
            (name, parent),
        )
        parent = int(cur.fetchone()[0])
    return parent


def parsed_brand(row: ImportRow) -> str:
    parsed = parse_product({"name": row.name, "category_path": ["Ваш тип товарів чи послуг", *row.category_path]})
    if not parsed:
        return row.manufacturer
    return parsed.brand or row.manufacturer


def upsert_product(conn, row: ImportRow, tip_id: int) -> int:
    cur = conn.cursor()
    existing = cur.execute("SELECT NUM FROM TOVAR_NAME WHERE KOD = ?", (row.code,)).fetchone()
    brand = parsed_brand(row)
    purchase = row.purchase_price or row.wholesale_price or row.retail_price

    params = (
        row.name,
        row.unit,
        tip_id,
        purchase,
        row.code,
        row.warranty,
        row.retail_price,
        row.wholesale_price,
        1,
        1,
        row.extra or row.source_file,
        int(row.packaging or 1),
        int(row.package_count or 1),
        row.min_qty,
        brand,
        row.promo_price,
    )

    if existing:
        cur.execute(
            """
            UPDATE TOVAR_NAME SET
                NAME = ?, ED_IZM = ?, TIP = ?, CENA = ?, VISIBLE = 1,
                KOD = ?, GARAN = ?, CENA_R = ?, CENA_O = ?,
                CENA_CURR_ID = ?, CENA_OUT_CURR_ID = ?, DOPOLN = ?,
                TOV_FASOVKA = ?, TOV_UPAKOVKA_COUNT = ?, KOLVO_MIN = ?,
                IS_USLUGA = 0, TOV_PROIZV = ?, CENA_PROMO = ?
            WHERE NUM = ?
            """,
            params + (int(existing[0]),),
        )
        return int(existing[0])

    cur.execute(
        """
        INSERT INTO TOVAR_NAME (
            NAME, ED_IZM, TIP, CENA, VISIBLE, KOD, GARAN, CENA_R, CENA_O,
            CENA_CURR_ID, CENA_OUT_CURR_ID, DOPOLN, IS_USLUGA,
            TOV_FASOVKA, TOV_UPAKOVKA_COUNT, KOLVO_MIN, TOV_PROIZV, CENA_PROMO
        )
        VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?, ?, ?, ?)
        RETURNING NUM
        """,
        params,
    )
    return int(cur.fetchone()[0])


def upsert_stock(conn, tovar_id: int, row: ImportRow) -> None:
    cur = conn.cursor()
    existing = cur.execute(
        """
        SELECT NUM FROM TOVAR_ZAL
        WHERE TOVAR_ID = ? AND FIRMA_ID = 1 AND SKLAD_ID = 1 AND VISIBLE = 1
        """,
        (tovar_id,),
    ).fetchone()
    purchase = row.purchase_price or row.wholesale_price or row.retail_price
    stock_params = (
        row.qty,
        row.qty * purchase,
        purchase,
        row.retail_price,
        row.wholesale_price,
    )
    if existing:
        cur.execute(
            """
            UPDATE TOVAR_ZAL SET
                KOLVO = ?, SUMA = ?, CENA_IN = ?, CENA_R = ?, CENA_O = ?, VISIBLE = 1
            WHERE NUM = ?
            """,
            stock_params + (int(existing[0]),),
        )
        return

    cur.execute(
        """
        INSERT INTO TOVAR_ZAL (
            FIRMA_ID, TOVAR_ID, SKLAD_ID, KOLVO, SUMA, CENA_IN, CENA_R, CENA_O, VISIBLE
        )
        VALUES (1, ?, 1, ?, ?, ?, ?, ?, 1)
        """,
        (tovar_id,) + stock_params,
    )


def collect_rows(paths: list[Path]) -> list[ImportRow]:
    rows: list[ImportRow] = []
    for path in paths:
        rows.extend(iter_import_rows(path))
    return rows


def write_preview(rows: list[ImportRow]) -> None:
    PREVIEW_JSON.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW_JSON.write_text(
        json.dumps([asdict(row) for row in rows], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def resolve_paths(directory: Path | None, files: list[str]) -> list[Path]:
    paths: list[Path] = []
    if directory:
        paths.extend(sorted(directory.glob("*.xlsx")))
    paths.extend(Path(item) for item in files)
    deduped: dict[str, Path] = {}
    for path in paths:
        deduped[str(path.resolve())] = path
    return list(deduped.values())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", type=Path, default=None, help="Папка з xlsx-файлами")
    ap.add_argument("--file", action="append", default=[], help="Окремий xlsx-файл")
    ap.add_argument("--db", choices=["live", "snapshot"], default="live")
    ap.add_argument("--apply", action="store_true", help="Реально записати в УкрСклад")
    args = ap.parse_args()

    paths = resolve_paths(args.dir, args.file)
    rows = collect_rows(paths)
    write_preview(rows)

    summary = {
        "files": len(paths),
        "rows": len(rows),
        "categories": sorted({"/".join(row.category_path) for row in rows}),
        "preview_json": str(PREVIEW_JSON),
        "apply": args.apply,
        "db": args.db,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if not args.apply:
        return 0

    db_path = LIVE_DB if args.db == "live" else SNAPSHOT_DB
    conn = connect(db_path)
    created_or_updated = 0
    try:
        for row in rows:
            tip_id = ensure_tip(conn, row.category_path)
            tovar_id = upsert_product(conn, row, tip_id)
            upsert_stock(conn, tovar_id, row)
            created_or_updated += 1
        conn.commit()
    finally:
        conn.close()

    print(json.dumps({"written": created_or_updated, "db_path": str(db_path)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
