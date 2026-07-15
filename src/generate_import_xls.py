"""
Генерує horoshop_import.xlsx для імпорту через адмінку Horoshop.

Важливий нюанс:
- Horoshop коректно авто-мапить Excel лише тоді, коли заголовки збігаються
  з його внутрішніми назвами полів.
- За актуальною документацією платформи для багатомовних полів потрібно
  використовувати формат без пробілу перед індексом мови: "Назва(ua)",
  "Опис товару(ua)" тощо.
"""
from __future__ import annotations
import json, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

ROOT = Path(r"D:\FISH\fish-sync")
sys.path.insert(0, str(ROOT / "src"))

from horoshop_catalog import build_canonical_products, collect_param_headers, strip_html

OUT = ROOT / "public" / "horoshop_import.xlsx"

BASE_COLUMNS = [
    "Артикул",
    "Назва(ua)",
    "Назва модифікації(ua)",
    "Розділ",
    "Цена",
    "Валюта",
    "Відображати",
    "Наявність",
    "Бренд",
    "Опис товару(ua)",
]


def clean_cell_value(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value

def main() -> None:
    products = build_canonical_products()
    param_headers = collect_param_headers(products)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товари"

    header_cols = BASE_COLUMNS + [f"{name}(ua)" for name in param_headers]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(header_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(15, len(col_name) + 2)

    for row_idx, p in enumerate(products, 2):
        params: dict[str, str] = {
            item["name"]: item["value"]
            for item in (p.get("params") or [])
        }

        title = p.get("title", "")
        row_data = [
            p.get("article", ""),
            title,
            title,
            p.get("parent", ""),
            p.get("price", "") if p.get("price") else "",
            "UAH",
            "Да" if int(p.get("display_in_showcase", 1)) else "Нет",
            p.get("presence", "Немає в наявності"),
            p.get("brand", ""),
            strip_html(p.get("description", "")),
        ]
        # Add param columns
        for param_name in param_headers:
            row_data.append(params.get(param_name, ""))

        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=clean_cell_value(value))

    # Freeze header row
    ws.freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Записано товарів: {len(products)}")
    print(f"Файл: {OUT}")


if __name__ == "__main__":
    main()
