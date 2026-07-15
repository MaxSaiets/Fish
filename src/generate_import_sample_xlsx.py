"""
Генерує малий XLSX для безпечного preview-тесту legacy імпорту Horoshop.

Це не фінальний імпорт. Файл містить тільки 5 товарів у наявності, щоб
перевірити, чи legacy `pricelist.php` коректно бачить колонки XLSX.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(r"D:\FISH\fish-sync")
sys.path.insert(0, str(ROOT / "src"))

from generate_import_csv import BASE_COLUMNS, clean_cell
from generate_import_sample_csv import SAMPLE_SIZE, is_safe_sample_product
from horoshop_catalog import build_canonical_products, collect_param_headers, strip_html

OUT = ROOT / "public" / "horoshop_import_sample_5.xlsx"
REPORT = ROOT / "data" / "horoshop_import_sample_5_xlsx_report.json"


def clean_xlsx_value(value: object) -> object:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", clean_cell(value))
    return value


def main() -> None:
    products = build_canonical_products()
    param_headers = collect_param_headers(products)
    headers = BASE_COLUMNS + [f"{name}(ua)" for name in param_headers]
    sample = [product for product in products if is_safe_sample_product(product)][:SAMPLE_SIZE]

    if len(sample) < SAMPLE_SIZE:
        raise RuntimeError(f"Знайдено лише {len(sample)} товарів для safe sample")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товари"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = min(45, max(15, len(col_name) + 2))

    for row_idx, product in enumerate(sample, 2):
        params = {
            item["name"]: item["value"]
            for item in product.get("params", [])
            if item.get("name")
        }
        title = product.get("title", "")
        row = [
            product.get("article", ""),
            title,
            title,
            product.get("parent", ""),
            product.get("price", "") if product.get("price") else "",
            "UAH",
            "Да" if int(product.get("display_in_showcase", 1)) else "Нет",
            product.get("presence", "Немає в наявності"),
            product.get("brand", ""),
            strip_html(product.get("description", "")),
        ]
        row.extend(params.get(param_name, "") for param_name in param_headers)
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=clean_xlsx_value(value))

    ws.freeze_panes = "A2"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    report = {
        "file": str(OUT),
        "rows": len(sample),
        "columns": len(headers),
        "purpose": "safe legacy import preview only, not final mass import",
        "products": [
            {
                "article": product.get("article", ""),
                "title": product.get("title", ""),
                "category": product.get("parent", ""),
                "price": product.get("price", ""),
                "presence": product.get("presence", ""),
                "params_count": len(product.get("params") or []),
            }
            for product in sample
        ],
    }
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
